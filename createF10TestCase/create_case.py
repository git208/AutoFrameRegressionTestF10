import openpyxl
from common.parse_excel import ParseExcel
import json
from createF10TestCase.market_stock_return import availableStock

class CreateCase(ParseExcel):

    def __init__(self,file,**kwargs):
        super().__init__(file,kwargs)
        self.wb = openpyxl.Workbook()
        for _ in self.wb.sheetnames:
            self.wb.remove(self.wb[_])

    def get_excel(self):
        wb = openpyxl.load_workbook(self.filepath)
        if self.sheet_name is None:
            ws = wb.active
        else:
            ws = wb[self.sheet_name]
        head_tuple = tuple(ws.iter_rows(max_row=1, values_only=True))[0]
        case_list = []
        for other_tuple in tuple(ws.iter_rows(min_row=2, values_only=True)):
            case_list.append(dict(zip(head_tuple, other_tuple)))
        return case_list

    def createCaseForF10(self):
        path_stocks_dict = {}
        for param_set in self.get_excel():
            if param_set['interface_abbreviation'] == None or param_set['interface_name'] == None:
                continue

            print(f'开始创建用例：{param_set["interface_name"]}')
            if param_set['interface_name'] in self.wb.sheetnames:
                self.ws = self.wb[param_set['interface_name']]
            else:
                self.ws = self.wb.create_sheet(param_set['interface_name'])

            self.ws['a1'] = 'test_case_id'
            self.ws['b1'] = 'test_case_name'
            self.ws['c1'] = 'path'
            self.ws['d1'] = 'url'
            self.ws['e1'] = 'method'
            self.ws['f1'] = 'body_type'
            self.ws['g1'] = 'headers'
            self.ws['h1'] = 'params'
            self.ws['i1'] = 'body'
            self.ws['j1'] = 'market'

            if param_set['auto_params'] == None or param_set['market'] == None:
                # print(param_set)
                if param_set['auto_param_structure'] == None:
                    max_row = self.ws.max_row
                    self.ws[f'a{max_row+1}'] = param_set['interface_abbreviation']+'_'+str(max_row+1)
                    self.ws[f'b{max_row+1}'] = param_set['interface_name']
                    self.ws[f'c{max_row+1}'] = param_set['path']
                    self.ws[f'e{max_row+1}'] = 'get'
                    self.ws[f'g{max_row+1}'] = param_set['headers']
                else:
                    auto_list = []
                    for key,value in json.loads(param_set['auto_param_structure']).items():
                        value_list = []
                        if type(value[0]) == str:
                            for _ in value:
                                value_list.append({key:_})
                            auto_list.append(value_list)
                        else:
                            temp_list01 = []
                            for _ in value:
                                temp_list02 = []
                                for __ in _:
                                    if temp_list01 == []:
                                        temp_list02.append(__)
                                    else:
                                        for ___ in temp_list01:
                                            temp_list02.append(str(___)+str(__))
                                temp_list01 = temp_list02
                            for _ in temp_list01:
                                value_list.append({key:_})
                            auto_list.append(value_list)

                    print(json.dumps(auto_list,ensure_ascii=False,indent=2))
                    auto_end_result_list = []
                    for _ in auto_list:
                        temp_list002 = []
                        for __ in _:
                            if auto_end_result_list == []:
                                temp_list002.append(__)
                            else:
                                for ___ in auto_end_result_list:
                                    dic = {}
                                    dic.update(___)
                                    dic.update(__)
                                    temp_list002.append(dic)
                        auto_end_result_list = temp_list002

                    i = 0
                    max_row = self.ws.max_row
                    for _ in auto_end_result_list:
                        i += 1
                        self.ws[f'a{max_row + i}'] = param_set['interface_abbreviation'] + '_' + str(max_row + i)
                        self.ws[f'b{max_row + i}'] = param_set['interface_name']
                        self.ws[f'c{max_row + i}'] = param_set['path']
                        self.ws[f'e{max_row + i}'] = 'get'
                        headers = json.loads(param_set['headers'])
                        headers.update(_)
                        print(f'总共{len(auto_end_result_list)}行，正在写入第{i}行，请求头为：{headers}')
                        self.ws[f'g{max_row + i}'] = json.dumps(headers)
            else:
                if param_set['auto_param_structure'] == None:

                    for _ in param_set['market'].split(','):
                        if param_set['path'] in path_stocks_dict.keys():
                            if _ in path_stocks_dict[param_set['path']].keys():
                                auto_stockdic_list = path_stocks_dict[param_set['path']][_]
                            else:
                                auto_stockdic_list = availableStock(param_set['path'], param_set['headers'],
                                                                    param_set['auto_params'], _,
                                                                    param_set['stock_search_site'])
                                path_stocks_dict.update({param_set['path']: {_: auto_stockdic_list}})
                        else:
                            auto_stockdic_list = availableStock(param_set['path'], param_set['headers'],
                                                                param_set['auto_params'], _ ,param_set['stock_search_site'])
                            path_stocks_dict.update({param_set['path']:{_:auto_stockdic_list}})

                        i = 0
                        max_row = self.ws.max_row
                        for __ in auto_stockdic_list:
                            i += 1
                            self.ws[f'a{max_row + i}'] = param_set['interface_abbreviation'] + '_' + str(max_row + i)
                            self.ws[f'b{max_row + i}'] = param_set['interface_name'] + '_' + __[param_set['auto_params']]
                            self.ws[f'c{max_row + i}'] = param_set['path']
                            self.ws[f'e{max_row + i}'] = 'get'
                            headers = json.loads(param_set['headers'])
                            headers.update(__)
                            print(f'总共{len(auto_stockdic_list)}行，正在写入第{i}行,数据：{headers}')
                            self.ws[f'g{max_row + i}'] = json.dumps(headers)
                            self.ws[f'j{max_row + i}'] = _
                else:
                    auto_list = []
                    for key, value in json.loads(param_set['auto_param_structure']).items():
                        value_list = []
                        if type(value[0]) == str:
                            for _ in value:
                                value_list.append({key: _})
                            auto_list.append(value_list)
                        else:
                            temp_list01 = []
                            for _ in value:
                                temp_list02 = []
                                for __ in _:
                                    if temp_list01 == []:
                                        temp_list02.append(__)
                                    else:
                                        for ___ in temp_list01:
                                            temp_list02.append(str(___) + str(__))
                                temp_list01 = temp_list02
                            for _ in temp_list01:
                                value_list.append({key: _})
                            auto_list.append(value_list)

                    auto_end_result_list = []
                    for _ in auto_list:
                        temp_list002 = []
                        for __ in _:
                            if auto_end_result_list == []:
                                temp_list002.append(__)
                            else:
                                for ___ in auto_end_result_list:
                                    dic = {}
                                    dic.update(___)
                                    dic.update(__)
                                    temp_list002.append(dic)
                        auto_end_result_list = temp_list002


                    for _ in param_set['market'].split(','):
                        if param_set['path'] in path_stocks_dict.keys():
                            if _ in path_stocks_dict[param_set['path']].keys():
                                auto_stockdic_list = path_stocks_dict[param_set['path']][_]
                            else:
                                auto_stockdic_list = availableStock(param_set['path'], param_set['headers'],
                                                                    param_set['auto_params'], _,
                                                                    param_set['stock_search_site'])
                                path_stocks_dict.update({param_set['path']: {_: auto_stockdic_list}})
                        else:
                            auto_stockdic_list = availableStock(param_set['path'], param_set['headers'],
                                                                param_set['auto_params'], _,
                                                                param_set['stock_search_site'])
                            path_stocks_dict.update({param_set['path']: {_: auto_stockdic_list}})

                        i = 0
                        max_row = self.ws.max_row
                        for __ in auto_stockdic_list:
                            for ___ in auto_end_result_list:
                                i += 1
                                self.ws[f'a{max_row + i}'] = param_set['interface_abbreviation'] + '_' + str(
                                    max_row + i)
                                self.ws[f'b{max_row + i}'] = param_set['interface_name'] + '_' + __[
                                    param_set['auto_params']]
                                self.ws[f'c{max_row + i}'] = param_set['path']
                                self.ws[f'e{max_row + i}'] = 'get'
                                headers = json.loads(param_set['headers'])
                                headers.update(___)
                                headers.update(__)
                                print(f'总共{len(auto_stockdic_list)* len(auto_end_result_list)}行，正在写入第{i}行,数据：{headers}')
                                self.ws[f'g{max_row + i}'] = json.dumps(headers)
                                self.ws[f'j{max_row + i}'] = _

            # self.wb.create_sheet('你是傻吊')
            # self.ws = self.wb['你是傻吊']
            # self.ws['C6'] = 'lalala'
        self.wb.save('../testCase/excels/F10用例(自动生成).xlsx')


if __name__ == '__main__':
    CreateCase('jiachao.xlsx').createCaseForF10()
