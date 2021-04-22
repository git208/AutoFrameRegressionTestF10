import openpyxl
import json
import time

import requests
class ParseExcel:

    def __init__(self,
                 file,
                 case_identifier=None,
                 sheet_name=None,
                 use_case_id=False
                 ):
        """
            :param file: 用例所在的路径
            :param case_identifier: 用例标识，为用例ID或用例序号
            :param sheet_name: 用例所在的sheet页
            :param use_case_id: 是否使用用例ID来进行驱动，置为False将使用用例序号进行驱动，默认使用用例序号进行驱动
        """
        self.data = None
        self.filepath = file
        self.sheet_name = sheet_name
        self.case_identifier = case_identifier
        self.use_case_id = use_case_id
        self.excel_data = {}

    def get_excel(self):
        """
        获取excel中的用例
        :return:
        """
        wb = openpyxl.load_workbook(self.filepath)

        if self.sheet_name is None:
            ws = wb.active
        else:
            ws = wb[self.sheet_name]
        head_tuple = tuple(ws.iter_rows(max_row=1, values_only=True))[0]
        case_list = []
        case_dic = {}
        for other_tuple in tuple(ws.iter_rows(min_row=2, values_only=True)):
            case_list.append(dict(zip(head_tuple, other_tuple)))
        for _ in case_list:
            case_dic[_['test_case_id']] = _
        return case_list,case_dic

    def get_excel_new(self):
        """
        获取excel中的用例
        :return:
        """
        wb = openpyxl.load_workbook(self.filepath)
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            head_tuple = tuple(ws.iter_rows(max_row=1, values_only=True))[0]
            case_list = []
            case_dic = {}
            for other_tuple in tuple(ws.iter_rows(min_row=2, values_only=True)):
                case_list.append(dict(zip(head_tuple, other_tuple)))
            for _ in case_list:
                case_dic[_['test_case_id']] = _
            self.excel_data.update({sheetname:[case_list, case_dic]})
        return self.excel_data

    def get_data(self):
        if self.excel_data == {}:
            self.get_excel_new()

        if self.use_case_id == False:
            if self.case_identifier == None:
                self.data = self.excel_data[self.sheet_name][0][0]
            else:
                self.data = self.excel_data[self.sheet_name][0][self.case_identifier - 2]
        else:
            if self.case_identifier != None:
                self.data = self.excel_data[self.sheet_name][1][self.case_identifier]
            else:
                self.data = {}

    def get_sheetnames(self):
        wb = openpyxl.load_workbook(self.filepath)
        return wb.sheetnames

    def data_process(self, key):
        self.get_data()
        if key in self.data.keys():
            if self.data[key] == "":
                return None
            else:
                return self.data[key]
        else:
            return None

    def get_test_case_id(self):
        return self.data_process('test_case_id')

    def get_test_case_name(self):
        return self.data_process('test_case_name')

    def get_url(self):
        return self.data_process('url')

    def get_path(self):
        return self.data_process('path')

    def get_method(self):
        return self.data_process('method')

    def get_type(self):
        return self.data_process('body_type')

    def get_headers(self):
        _ = self.data_process('headers')
        return json.loads(_) if _ != None else None

    def get_params(self):
        _ = self.data_process('params')
        return json.loads(_) if _ != None else None

    # def get_auto_params(self):
    #     _ = self.data_process('auto_params')
    #     return json.loads(_) if _ != None else None
    #
    # def get_market(self):
    #     _ = self.data_process('market')
    #     return json.loads(_) if _ != None else None
    #
    # def get_interface_name(self):
    #     _ = self.data_process('interface_name')
    #     return json.loads(_) if _ != None else None
    #
    # def get_interface_abbreviation(self):
    #     _ = self.data_process('interface_abbreviation')
    #     return json.loads(_) if _ != None else None

    def get_body(self):
        try:
            _ = self.get_type().lower()
        except:
            return None
        else:
            __ = self.data_process('body')
            if _ == 'form-data' or _ == 'x-www-form-urlencoded' or _ == 'json':
                return json.loads(__) if _ != None else None
            else:
                return __


if __name__ == '__main__':
    # pass
    das = ParseExcel(file='../testCase/excels/接口.xlsx', case_identifier='GG_004', use_case_id=True)
    # print(das.get_excel()[1].keys())
    # print(json.dumps([i for i in das.get_excel()[1].keys()], ensure_ascii=False, indent=2))
    print(das.get_sheetnames())
    # print(das.get_test_case_id())

